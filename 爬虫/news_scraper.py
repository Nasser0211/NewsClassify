# encoding=utf-8

import requests
import json
from lxml import etree
import openpyxl


class NewsScraper(object):
    # 所有分类
    CATEGORY = {
        '财经': 'finance',
        '游戏': 'games',
        '体育': 'sports',
        '娱乐': 'ent',
        '科技': 'tech',
        '汽车': 'auto',
        '房产': 'house',
        '军事': 'milite'
    }

    # api
    API = 'https://i.news.qq.com/trpc.qqnews_web.kv_srv.kv_srv_http_proxy/list'

    # 请求参数公共部分
    params = {
        "srv_id": "pc",
        'strategy': 1,
        'ext': {'pool': ['top'], 'is_filter': 2, 'check_type': True},
        'offset': 0,
        'limit': 199,
        'sub_srv_id': "milite",
    }

    def __init__(self, file='源数据2.xlsx'):
        self.file = file
        self.file_is_existed = False
        # 将参数内部的对象序列化
        NewsScraper.params['ext'] = json.dumps(NewsScraper.params["ext"])

    # api → 新闻详情url列表
    def request_news_list(self, cate):

        # 1. 参数整理
        NewsScraper.params['sub_srv_id'] = NewsScraper.CATEGORY[cate]  # 获取分类

        # 2. 请求
        res = requests.get(url=NewsScraper.API, params=NewsScraper.params)

        # 3. 解析结果
        print(json.loads(res.text))
        print(res.url)
        news = json.loads(res.text)['data']['list']  # json解析为对象
        urls = [a_news['url'] for a_news in news]  # 提取详情url

        return urls

    # 新闻详情url(列表) →  新闻数据数组rows
    def request_news_detail(self, urls):
        rows = []
        xpath_title_str = "//div[@class='LEFT']/h1/text()"
        xpath_content_str = "//p[@class='one-p']/text()"

        # 1. 循环每个新闻url
        for url in urls:

            # 2. 请求新闻详情
            res = requests.get(url).text

            # 3. 解析结果
            html = etree.HTML(res)  # 将新闻内容字符串转化为特殊对象
            # 调用方法,解析,获得需要的值
            if not html.xpath(xpath_title_str):
                continue
            if not html.xpath(xpath_content_str):
                continue
            title = html.xpath(xpath_title_str)[0]
            content = html.xpath(xpath_content_str)

            # 每段去掉空格
            content_clean = [p.strip() for p in content]
            content = "".join(content_clean)

            rows.append({"title": title, "content": content})
        return rows

    # 数据数组rows  →  保存
    def save_data(self, sheet=None, rows=None):
        # 判断工作簿是否存在
        if self.file_is_existed:
            print("文件已存在----------------------------------------")
            workbook = openpyxl.load_workbook(self.file)
        else:
            print("文件不存在----------------------------------------")
            workbook = openpyxl.Workbook()
            st = workbook.active
            st.title = sheet


        # 判断工作表是否存在
        print(sheet)
        print(workbook.sheetnames)
        if sheet not in workbook.sheetnames:
            print("工作表不存在，新建工作表——————————————————————————————————————")
            sheet = workbook.create_sheet(sheet)
        else:
            sheet = workbook[sheet]

        # 循环所有新闻，添加数据
        for news in rows:
            sheet.append([news['title'], news['content']])

        # 保存
        workbook.save(self.file)
        self.file_is_existed = True
        return

    def turn_up(self):
        # 循环所有分类
        for cate in NewsScraper.CATEGORY:
            print("开始收集" + cate + "的新闻-------------------------------------------")
            urls = news_scraper.request_news_list(cate)
            rows = news_scraper.request_news_detail(urls)
            news_scraper.save_data(cate, rows)

# 创建类，  无参数，默认的保存路径为 源数据2.xlsx
news_scraper = NewsScraper()
news_scraper.turn_up()
